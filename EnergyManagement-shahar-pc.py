# -*- coding: utf-8 -*-
import csv
import numpy as np
import os.path
import pickle
from datetime import datetime, timedelta
from pprint import pprint
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt 
from scipy.interpolate import CubicSpline
from scipy import optimize

from tensorforce.environments import Environment
from tensorforce.agents import PPOAgent
from tensorforce.execution import Runner

# run parameters
# ==============

DATA_IO = 'save' # 'save' or 'load'

if DATA_IO == 'save':
  # read electricity data
  # =====================
  
  # prepare storage
  elect = []
  
  # read data
  print('reading electricity file...')
  fname = './data/Electricity Data.xlsx'
  wb = load_workbook(filename=fname, read_only=True)
  ws = wb['Electricity Data']
  for row in ws.iter_rows(min_row=4, max_row=8763):
    elect.append([row[0].value, row[2].value, row[3].value, row[6].value])
  
  # fix problem - time 00:00 should advance one day
  for row in elect[1:]:
    if row[0].hour==0:
      row[0] += timedelta(days=1)
  
  # elect content:
  #  [datatime, demand [MW], solar-generation [MW], marginal cost[ag/kWh]] - every 3 hours
  
  pprint(elect[:5])
  
  x = [r[0] for r in elect]
  dmnd = np.array([r[1] for r in elect])
  renw = np.array([r[2] for r in elect])
  cost = np.array([r[3] for r in elect])
  
  #plt.plot(x[:150], dmnd[:150], 'demand', x[:150], renw[:150], 'renewables', x[:150], cost[:150], 'cost')
  plt.plot(x[:150], dmnd[:150], label='demand')
  plt.plot(x[:150], renw[:150], label='renewables')
  plt.plot(x[:150], cost[:150], label='cost')
  plt.title('electricity data')
  plt.ylabel('MW')
  plt.legend(loc=2)
  plt.show()
  
  
  # read meteorological data
  # ========================
  
  # prepare storage
  meteo = []
  
  # read data
  print('reading meteorology file...')
  with open("./data/טמפ' תל-אביב חוף 2017.csv", 'r', encoding="utf8") as f:
    reader = csv.reader(f)
    firstRow = reader.__next__()
    print('headers:', firstRow)
  
    for row in reader:
      dt = datetime.strptime('{} {}'.format(row[2], row[3]), '%d-%m-%Y %H:%M')
      temp = float(row[4])
      meteo.append([dt, temp])
  
  #meteo = np.array(meteo)
  
  # meteo content:
  #  [datatime, temp] - every 3 hours
  
  pprint(meteo[:5])
  
  x = [r[0] for r in meteo]
  y = [r[1] for r in meteo]
  
  plt.plot(x[:50], y[:50], 'o', x[:50], y[:50], '-')
  plt.title('temperature - 3-hours data')
  plt.ylabel('temperature')
  plt.show()
  
  # use x timestamp since its integer
  xts = [r[0].timestamp() for r in meteo]
  
  # set x1 from the hours in elect array
  x1 = [r[0] for r in elect]
  
  # convert to timestamp, for interpolation
  x1ts = [xx.timestamp() for xx in x1]
  
  # interpolate hourly meteo using cuibic spline interpolation
  cs = CubicSpline(xts, y)
  y1 = cs(x1ts)
  
  plt.plot(x1[:150], y1[:150], 'o', x1[:150], y1[:150], '-')
  plt.title('temperature - 1-hours data')
  plt.ylabel('temperature')
  plt.show()
  
  # save data
  outfile = open('./data/input.dat', 'wb')
  pickle.dump(x1, outfile)
  pickle.dump(y1, outfile)
  pickle.dump(dmnd, outfile)
  pickle.dump(renw, outfile)
  pickle.dump(cost, outfile)
  outfile.close()

else:
  # load data
  infile = open('./data/input.dat', 'rb')
  x1 = pickle.load(infile)
  y1 = pickle.load(infile)
  dmnd = pickle.load(infile)
  renw = pickle.load(infile)
  cost = pickle.load(infile)
  infile.close()


# manual handle batteries
# =====================
battMaxCharge = 3500 # 3.5 TWh
battState = 0
batt = np.array(dmnd.size)
# save dmnd data
dmndOrig = np.copy(dmnd)

for day in range(365):
  # find min points in current and next day
  dayStart = day*24
  todaysDemand = dmnd[dayStart : dayStart + 24]
  todaysMinPos = dayStart + np.argmin(todaysDemand)
  if day < 364:
    # regular day: find the minimum demand on the next day. usually at about
    # 01:00-04:00 on the next day's morning
    tomorrowsDemand = dmnd[dayStart + 24 : dayStart + 48]
    tomorrowsMinPos = dayStart + 24 + np.argmin(tomorrowsDemand)
  else:
    # on the last day of the year, can't look for minimum on the next day,
    # so take the last hour of the day. it's good enough
    tomorrowsMinPos = 365*24
  
  # find total solar production
  todaysDemand = dmnd[todaysMinPos : tomorrowsMinPos]
  todaysSolar = renw[todaysMinPos : tomorrowsMinPos]
  totalSolarProduction = todaysSolar.sum()
  # cap battery charge by max battery capacity
  battCharge = min(totalSolarProduction, battMaxCharge - battState)
  # charge2gridRatio is the ratio between solar power for charging and power for grid use
  charge2gridRatio = battCharge / totalSolarProduction
  solarToGrid = todaysSolar * (1-charge2gridRatio)
  
  # find minimal max fossil power to use when discharging battery
  # first, subtract the solar power that is routed to the grid from todaysDemand
  todaysDemand -= solarToGrid
  maxDemand = todaysDemand.max()
  
  # dischargeBattFun(maxDemand) returns the max power demand when discharging the battery
  # above maxDemand
  def dischargeBattFun(maxDemand, returnAllData = False):
    curCharge = battState
    todaysBattState = np.zeros(todaysDemand.size)
    for h in range(todaysDemand.size):
      todaysBattState[h] = battState
      # charge the battery from solar
      curCharge += todaysSolar[h]*charge2gridRatio
      curCharge = min(curCharge, battMaxCharge)
      # if demand is higher than masDemand, start discharging the battery
      if todaysDemand[h] > maxDemand:
        # calc discharge amount
        toDischarge = todaysDemand[h] - maxDemand
        toDischarge = min(toDischarge, curCharge)
        # reduce the demand by the discharge
        todaysDemand[h] -= toDischarge
        # discharge the battery
        curCharge -= toDischarge
    # return short or long format
    if not returnAllData:
      return todaysDemand.max()
    return todaysDemand.max(), todaysDemand, todaysBattState
  
  # use dischargeBattFun to find the minimal possible maxDemand
  minMaxDemand = optimize.fmin(dischargeBattFun, maxDemand * 0.9)
  
  # put updated values in arrays
  minMaxDemand, todaysDemand, todaysBattState = dischargeBattFun(maxDemand, returnAllData = True)
  dmnd[dayStart : dayStart + 24] = todaysDemand
  batt[dayStart : dayStart + 24] = todaysBattState
  battState = todaysBattState[-1]
  
# save updated values
fname = './data/Results.xlsx'
if not os.path.isfile(fname):
  # create workbook
  wb = openpyxl.Workbook()
  ws = wb.worksheets[0]
  ws.title = 'Results'
  ws['A1'] = 'datetime'
  ws['B1'] = 'demand-orig'
  ws['C1'] = 'demand-mod'
  ws['D1'] = 'renewables'
  ws['E1'] = 'battery'  
  wb.save(fname)
  
wb = load_workbook(filename=fname, read_only=False)
ws = wb.worksheets[0]
for line in range(dmnd.size):
  ws.cell(row=line+2, column=1).value = x[line]
  ws.cell(row=line+2, column=2).value = dmndOrig[line]
  ws.cell(row=line+2, column=3).value = dmnd[line]
  ws.cell(row=line+2, column=4).value = renw[line]
  ws.cell(row=line+2, column=5).value = batt[line]

wb.save()
