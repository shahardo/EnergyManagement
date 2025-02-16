# -*- coding: utf-8 -*-
import csv
import numpy as np
import os.path
import pickle
import random
import time
from datetime import datetime, timedelta
from pprint import pprint
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from scipy.interpolate import CubicSpline

# # set IPython charts inline (use 'qt' for a new Qt window)
# from IPython import get_ipython
# get_ipython().run_line_magic('matplotlib', 'inline')

# run parameters
# ==============

DATA_IO = 'load' # 'save' or 'load'
PROCESS = 'rl' # 'rl' or 'manual'
DEBUG_CHARTS = False
NUMBER_OF_RUNS = 'single' # 'single' or 'multiple'

# prepare run parameters

if NUMBER_OF_RUNS == 'single':
  battMaxCharge = 3500 # 3.5 TWh
  targetRenewable = 40 #%
else:
  battMaxChargeValues = [0, 1000, 2000, 3000, 3500, 4000] # MWh
  targetRenewableValues = [0, 10, 20, 30, 40, 50, 100] #%

if DATA_IO == 'save':

  # read electricity data
  # =====================

  # prepare storage
  elect = []

  # read data
  print('reading electricity file...')
  x1 = []
  dmnd = []
  renw = []
  batt = []

  fname = './data/Results - 40pcnt renew, 3500 storage.xlsx'
  wb = load_workbook(filename=fname, read_only=True)
  ws = wb['Results']
  for row in ws.iter_rows(min_row=4, max_row=8763):
    x1.append(row[0].value)
    dmnd.append(row[2].value)
    renw.append(row[4].value)
    batt.append(row[5].value)

  x1 = np.array(x1)
  dmnd = np.array(dmnd)
  renw = np.array(renw)
  batt = np.array(batt)

  # pprint([[x1[i], dmnd[i], renw[i], batt[i]] for i in range(5)])

  #plt.plot(x[:150], dmnd[:150], 'demand', x[:150], renw[:150], 'renewables', x[:150], cost[:150], 'cost')
  plt.plot(x1[:150], dmnd[:150], label='demand')
  plt.plot(x1[:150], renw[:150], label='renewables')
  plt.plot(x1[:150], batt[:150], label='battery')
  plt.title('electricity data')
  plt.ylabel('MW')
  plt.legend(loc=2)
  plt.show()


  # read meteorological data
  # ========================

  # prepare storage
  meteo = []

  # read data
  print('reading meteorology file...')
  with open("./data/טמפ' תל-אביב חוף 2017.csv", 'r', encoding="utf8") as f:
    reader = csv.reader(f)
    firstRow = reader.__next__()
    print('headers:', firstRow)

    for row in reader:
      dt = datetime.strptime('{} {}'.format(row[2], row[3]), '%d-%m-%Y %H:%M')
      temp = float(row[4])
      meteo.append([dt, temp])

  #meteo = np.array(meteo)

  # meteo content:
  #  [datatime, temp] - every 3 hours

  pprint(meteo[:5])

  x = [r[0] for r in meteo]
  y = [r[1] for r in meteo]

  plt.plot(x[:50], y[:50], 'o', x[:50], y[:50], '-')
  plt.title('temperature - 3-hours data')
  plt.ylabel('temperature')
  plt.show()

  # use x timestamp since its integer
  xts = [r[0].timestamp() for r in meteo]

  # convert date-time to timestamp, for interpolation
  x1ts = [xx.timestamp() for xx in x1]

  # interpolate hourly meteo using cuibic spline interpolation
  cs = CubicSpline(xts, y)
  y1 = cs(x1ts)

  plt.plot(x1[:150], y1[:150], 'o', x1[:150], y1[:150], '-')
  plt.title('temperature - 1-hours data')
  plt.ylabel('temperature')
  plt.show()

  # save data
  print('saving input data')
  outfile = open('./data/input.dat', 'wb')
  pickle.dump(x1, outfile)
  pickle.dump(y1, outfile)
  pickle.dump(dmnd, outfile)
  pickle.dump(renw, outfile)
  pickle.dump(batt, outfile)
  outfile.close()

else:
  # load data
  print('reading data')
  infile = open('./data/input.dat', 'rb')
  x1 = pickle.load(infile)
  y1 = pickle.load(infile)
  dmnd = pickle.load(infile)
  renw = pickle.load(infile)
  batt = pickle.load(infile)
  infile.close()


if PROCESS == 'manual':
  # manual handle batteries
  # =======================
  from scipy import optimize

  def doManualOptimization(battMaxCharge, dmnd, renw):
    battState = 0
    batt = np.zeros(dmnd.size) # battery charge array
    wast = np.zeros(dmnd.size) # wasted energy array

    for day in range(365):
      # find min points in current and next day
      dayStart = day*24
      print(x1[dayStart].strftime('%d/%m/%Y'))
      todaysDemand = dmnd[dayStart : dayStart + 24]
      todaysMinPos = dayStart + np.argmin(todaysDemand)
      if day < 364:
        # regular day: find the minimum demand on the next day. usually at about
        # 01:00-04:00 on the next day's morning
        tomorrowsDemand = dmnd[dayStart + 24 : dayStart + 48]
        tomorrowsMinPos = dayStart + 24 + np.argmin(tomorrowsDemand)
      else:
        # on the last day of the year, can't look for minimum on the next day,
        # so take the last hour of the day. it's good enough
        tomorrowsMinPos = 365*24

      # find total solar production
      todaysDemand = dmnd[todaysMinPos : tomorrowsMinPos].copy()
      todaysSolar = renw[todaysMinPos : tomorrowsMinPos]
      totalSolarProduction = todaysSolar.sum()
      # cap battery charge by max battery capacity
      battCharge = min(totalSolarProduction, battMaxCharge - battState)
      # charge2gridRatio is the ratio between solar power for charging and power for grid use
      charge2gridRatio = battCharge / totalSolarProduction if totalSolarProduction != 0 else 0
      solarToGrid = todaysSolar * (1-charge2gridRatio)

      # find minimal max fossil power to use when discharging battery
      # first, subtract the solar power that is routed to the grid from todaysDemand
      todaysDemand -= solarToGrid
      todaysWasted = np.array([max(0, -d) for d in todaysDemand]) # if solarToGrid is higher than demand
      todaysDemand = np.array([max(0,  d) for d in todaysDemand]) # can't have negative demand

      # dischargeBattFun(targetDemand) returns the max power demand when discharging the battery
      # above targetDemand
      def dischargeBattFun(targetDemand, returnAllData = False):
        targetDemand = targetDemand[0] # targetDemand received as array
        todaysModifiedDemand = todaysDemand.copy() # so todaysDemand on the outer scope won't change
        curCharge = battState
        todaysBattState = np.zeros(todaysDemand.size)
        for h in range(todaysDemand.size):
          # charge the battery from solar
          curCharge += todaysSolar[h]*charge2gridRatio
          todaysWasted[h] += max(0, curCharge - battMaxCharge) # add amount of wasted solar energy
          curCharge = min(curCharge, battMaxCharge) # cap charge with battery max capacity
          # if demand is higher than maxDemand, start discharging the battery
          if todaysModifiedDemand[h] > targetDemand:
            # calc discharge amount
            toDischarge = todaysModifiedDemand[h] - targetDemand
            toDischarge = min(toDischarge, curCharge)
            # reduce the demand by the discharge
            todaysModifiedDemand[h] -= toDischarge
            # discharge the battery
            curCharge -= toDischarge
          # save battery charege state
          todaysBattState[h] = curCharge

        # debug reuslts
    #    plt.plot(todaysDemand, label='demand')
    #    plt.plot(todaysModifiedDemand, label='mod-demand')
    #    plt.plot(todaysSolar, label='renewables')
    #    plt.plot(todaysBattState, label='batt')
    #    plt.title('electricity data')
    #    plt.ylabel('MW')
    #    plt.legend(loc=2)
    #    plt.show()

        # return short or long format
        if not returnAllData:
          return todaysModifiedDemand.max()
        return todaysModifiedDemand.max(), todaysModifiedDemand, todaysBattState, todaysWasted

      # use dischargeBattFun to find the minimal possible maxDemand
      maxDemand = todaysDemand.max()
      minMaxDemand = optimize.fmin(dischargeBattFun, maxDemand * 0.9)

      # put updated values in arrays
      minMaxDemand, todaysModifiedDemand, todaysBattState, todaysWasted = dischargeBattFun([minMaxDemand], returnAllData = True)
      dmnd[todaysMinPos : todaysMinPos + todaysModifiedDemand.size] = todaysModifiedDemand
      batt[todaysMinPos : todaysMinPos + todaysModifiedDemand.size] = todaysBattState
      wast[todaysMinPos : todaysMinPos + todaysModifiedDemand.size] = todaysWasted
      battState = todaysBattState[-1]

      # debug reuslts
      if DEBUG_CHARTS:
        plt.plot(todaysDemand, label='demand')
        plt.plot(todaysModifiedDemand, label='mod-demand')
        plt.plot(todaysSolar, label='renewables')
        plt.plot(todaysBattState, label='batt')
        plt.plot(todaysWasted, label='wasted')
        plt.title(x1[dayStart].strftime('%d/%m/%Y'))
        plt.ylabel('MW')
        plt.legend(loc=2)
        plt.show()

    return dmnd, batt, wast

  # prepare output file
  curDate = datetime.today().strftime('%Y-%m-%d')
  fname = './data/Results - {}.xlsx'.format(curDate)
  if os.path.isfile(fname):
    os.remove(fname)

  # create workbook
  wb = openpyxl.Workbook()
  ws = wb.worksheets[0]
  ws.title = 'Results'
  ws['A1'] = 'datetime'
  ws['B1'] = 'hour'
  ws['C1'] = 'ren-part'
  ws['D1'] = 'storage'
  ws['E1'] = 'demand-orig'
  ws['F1'] = 'demand-mod'
  ws['G1'] = 'renewables'
  ws['H1'] = 'battery'
  ws['I1'] = 'waste'
  wb.save(fname)
  outLine = 2

  # loop for all ren/storage options
  for maxBat in battMaxChargeValues:
    for renPart in targetRenewableValues:
      print('running optimization')
      print('batt max charge: {}, target renewables: {}%'.format(maxBat, renPart))

      # adjust solar production for expected target
      renewableRatio = renw.sum() / dmnd.sum()
      renewableFactor = renPart / 100 / renewableRatio
      renwAdjusted = renw * renewableFactor

      # run optimization for a full year, according to batt-max-charte and target-renewables
      # use a copy of dmnd since it's going to change during optimization
      dmndMod, batt, wast = doManualOptimization(maxBat, np.copy(dmnd), renwAdjusted)

      print('saving results to excel')
#      wb = load_workbook(filename=fname, read_only=False)
#      ws = wb.worksheets[0]
      for i in range(dmnd.size):
        ws.cell(row=outLine, column=1).value = x1[i]
        ws.cell(row=outLine, column=3).value = renPart/100
        ws.cell(row=outLine, column=4).value = maxBat
        ws.cell(row=outLine, column=5).value = float(dmnd[i])
        ws.cell(row=outLine, column=6).value = float(dmndMod[i])
        ws.cell(row=outLine, column=7).value = float(renwAdjusted[i])
        ws.cell(row=outLine, column=8).value = float(batt[i])
        ws.cell(row=outLine, column=9).value = float(wast[i])
        outLine +=1

  wb.save(fname)
  print('done!')

if PROCESS == 'rl':
  # RL handle batteries
  # =======================
  print('importing tensorforce')
  from tensorforce.environments import Environment
  from tensorforce.agents import PPOAgent, DQFDAgent
  from tensorforce.execution import Runner


  #  normalize input
  # =======================
  def norm(x):
    # normalize data to have avg=0, scale=2 (+1..-1)
    scale = (x.max()-x.min())
    return x/scale

  dmnd = norm(dmnd)
  renw = norm(renw)
#  cost = norm(cost)
  batt = norm(batt)
  y1 = norm(y1)


  # prepare environment
  # =======================

  class EnergyEnvironment(Environment):
    """
    Energy environment class
    """
    def reset(self, day=None):
      """
      Reset env to random day in the year
      if day is given, take that day (used for import_deomonstrations)
      otherwise, use a random day
      """
      # reset battery charge state
      self.battCharge = 0

      # reset episode hour. will run for 24*14 hours
      self.episodeHour = 0

      # reset thrown electricity accumulator
      self.thrownElect = 0

      # start episode from random day
      # start from day 2 (to allow for 24 hours of past data)
      # end 2 weeks befor year end (to allow for 14 days episode)
      # hour will be the index to yearly consumption/meteorology arrays
      self.hour = day * 24 if day else random.randint(1, 365-14)*24

      # some constants
      self.battMaxCharge = 0.1 # max demand ~12,000. take 10% of that. was 1000
      self.episodeHours = 1*24 # 14 days * 24 hours
      self.electTariff = 0.1

      return self.currentState()

    # current env state
    def currentState(self):
      '''
      the state includes 73 floats:
        24 floats: prev 24 hours consumption
        24 floats: prev 24 hours temperature
        24 floats: next 24 hours temperature forcast
        1  float : current battery charge state
      '''
      pastCons = dmnd[self.hour-24 : self.hour]
      pastMeteo = y1[self.hour-24 : self.hour]
      futureMeteo = y1[self.hour : self.hour+24]
      return np.concatenate((pastCons, pastMeteo, futureMeteo, [self.battCharge]))

    # action is battery charge command
    # positive to charge, negative to discharge
    def execute(self, action):
      # make sure action is a scalar
      if isinstance(action, np.ndarray):
        assert(action.size==1)
        action = action[0]

      # calc reward (the payment for generated electricity)
      # renewables are positive, demand is negative
      netDemand = renw[self.hour] - dmnd[self.hour]
      # for negative net-demand, don't allow charging the battery
      if netDemand < 0: action = max(action, 0)

      # check charge limits
      # can't discharge below zero
      # can't charge above battMaxCharge
      if -action > self.battCharge: action = -self.battCharge
      if action + self.battCharge > self.battMaxCharge: action = self.battMaxCharge - self.battCharge

      # charge/discharge battery
      netDemand += -action
      self.battCharge += action
      print('{},'.format(action), end='') # TODO: DELME

      # collect thrown away ren elect
      if netDemand > 0:
        self.thrownElect += netDemand
        netDemand = 0

      # calc electricity cost
      # at this stage, netDemand is either negative (need to generate
      # with fossil means), or zero
      electCost = -netDemand * self.electTariff # cost[self.hour]

      # get terminal state
      self.hour += 1
      self.episodeHour += 1
      terminal = self.episodeHour >= self.episodeHours

      return self.currentState(), terminal, -electCost

    @property
    def states(self):
      """
      73 floats:
        24 floats: prev 24 hours consumption
        24 floats: prev 24 hours temperature
        24 floats: next 24 hours temperature forcast
        1  float : current battery charge state
      """
      return dict(type='float', shape=73)

    @property
    def actions(self):
      """
      1 float: battery charge command: positive to charge, negative for discharge
      """
      return dict(type='float', shape=1)

  print('creating environment')
  env = EnergyEnvironment()


  # Instantiate a Tensorforce agent
  # ===============================

  networkFirstLayer = env.states['shape']*2 # two times the states space - 2*73
  networkLastLayer = env.actions['shape']*10 # ten times the actions space - 10*1

  # prepare agent

  def createPPOAgent():
    agent = PPOAgent(
      states = env.states,
      actions = env.actions,
      network=[
        dict(type='dense', size=networkFirstLayer),
        dict(type='dense', size=int((networkFirstLayer*networkLastLayer)**0.5)), # geometric average of first and last
        dict(type='dense', size=networkLastLayer),
      ],
      step_optimizer=dict(type='adam', learning_rate=1e-2)
    )
    return agent

  def createPPO2Agent():
    # based on: https://github.com/tensorforce/tensorforce/blob/master/examples/quickstart.py
    agent = PPOAgent(
      states=env.states,
      actions=env.actions,
      network=[
        dict(type='dense', size=64),
        dict(type='dense', size=32)
      ],
      # Agent
      states_preprocessing=None,
      actions_exploration=None,
      reward_preprocessing=None,
      # MemoryModel
      update_mode=dict(
          unit='episodes',
          # 10 episodes per update
          batch_size=10,
          # Every 10 episodes
          frequency=10
      ),
      memory=dict(
          type='latest',
          include_next_states=False,
          capacity=5000
      ),
      # DistributionModel
      distributions=None,
      entropy_regularization=0.01,
      # PGModel
      baseline_mode='states',
      baseline=dict(
          type='mlp',
          sizes=[32, 32]
      ),
      baseline_optimizer=dict(
          type='multi_step',
          optimizer=dict(
              type='adam',
              learning_rate=1e-3
          ),
          num_steps=5
      ),
      gae_lambda=0.97,
      # PGLRModel
      likelihood_ratio_clipping=0.2,
      # PPOAgent
      step_optimizer=dict(
          type='adam',
          learning_rate=1e-3
      ),
      subsampling_fraction=0.2,
      optimization_steps=25,
      execution=dict(
          type='single',
          session_config=None,
          distributed_spec=None
      )
    )
    return agent

  def createDQFDAgent(states, actions, rewards, terminals):
    agent = DQFDAgent(
      states = env.states,
      actions = env.actions,
      network=[
        dict(type='dense', size=networkFirstLayer),
        dict(type='dense', size=int((networkFirstLayer*networkLastLayer)**0.5)), # geometric average of first and last
        dict(type='dense', size=networkLastLayer),
      ],
      optimizer=dict(type='adam', learning_rate=1e-4),
      target_sync_frequency=10,
    )
    demonstrations = dict(
      states = states,
      #internals = internals,
      actions = actions,
      terminal = terminal,
      reward = reward
    )
    agent.import_demonstrations(demonstrations = demonstrations)
    agent.pretrain(steps = 24 * 10)
    return agent


  # create agent and feed demonstration data
  # =========================================

  states = []
  actions = []
  terminals = []
  rewards = []

  AGENT_TYPE = 'ppo'

  if 'dqfd' == AGENT_TYPE:
    print('preparing demonstration data')
    for day in range(1, 356-7):
      env.reset(day)
      terminal = False
      while not terminal:
        action = batt[day] - batt[day-1]
        state, terminal, reward = env.execute(action)
        if terminal: print()
        actions.append(action)
        states.append(state)
        rewards.append(reward)
        terminals.append(terminal)
        day += 1

  # create agent and feed it with demonstration data
  print('creating agent ({})'.format(AGENT_TYPE))
  if 'dqfd' == AGENT_TYPE:
    agent = createDQFDAgent(states, actions, rewards, terminals)
  elif 'ppo' == AGENT_TYPE:
    agent = createPPOAgent()
  elif 'ppo2' == AGENT_TYPE:
    agent = createPPO2Agent()
  else:
    assert('bad agent type')


  # Create the runner
  # =======================

  print('creating runner')
  history = {}
  runner = Runner(agent=agent, environment=env, history=history)

  # if False:
  #   # open output chart window
  #   # create figure (will only create new window if needed)
  #   plt.figure()
  #   # Generate plot1
  #   plt.plot(range(10, 20))
  #   # Show the plot in non-blocking mode
  #   plt.show(block=False)

#  fig, ax = plt.subplots()
#  ax.plot(range(5))
#  fig.canvas.manager.show()
#  # this makes sure that the gui window gets shown
#  # if this is needed depends on rcparams, this is just to be safe
#  fig.canvas.flush_events()
#  # this make sure that if the event loop integration is not
#  # set up by the gui framework the plot will update
#  time.sleep(0.5)
#  fig.canvas.flush_events()

  # moving average helper fun
  window_width = 10
  def movingAverage(data):
    cumsum_vec = np.cumsum(np.insert(data, 0, 0))
    return (cumsum_vec[window_width:] - cumsum_vec[:-window_width]) / window_width

  rewares = []

  # Callback function printing episode statistics
  def episode_finished(r):
    print("\nFinished episode {ep} after {ts} timesteps (reward: {reward})".format(
        ep=r.episode,
        ts=r.episode_timestep,
        reward=r.episode_rewards[-1])
    )
    if r.episode%10==0:
      rewards.append(r.episode_rewards[-1])
#    ax.plot(rewards)
#    fig.canvas.manager.show()
#    fig.canvas.flush_events()
    if r.episode%50==0:
      plt.figure()
      plt.plot(rewards)
      plt.plot(movingAverage(rewards))
      plt.show()

    return True

  # Start learning
  runner.run(episodes=100000, max_episode_timesteps=400, episode_finished=episode_finished)#, deterministic=True)
  runner.close()

  # Print statistics
  print("Learning finished. Total episodes: {ep}. Average reward of last 100 episodes: {ar}.".format(
      ep=runner.episode,
      ar=np.mean(runner.episode_rewards[-100:]))
  )